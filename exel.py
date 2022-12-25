import openpyxl
book = openpyxl.load_workbook('ВСЕ БАЗЫ.xlsx')
sheet = book.active
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
wb_ = Workbook()
ws_ = wb_.active
cells = sheet['A1':'E1']


for a,b,c,d,e in cells:
    A1 = a.value
    B1 = b.value
    C1 = c.value
    D1 = d.value
    E1 = e.value
    ws['A1'] = A1
    ws['B1'] = B1
    ws['C1'] = C1
    ws['D1'] = D1
    ws['E1'] = E1

wb.save('new file.xlsx')

r = 2
for row in range(2, 15):
# for row in range(2, sheet.max_row+1):

    data = sheet[row][3].value
    baza = sheet[row][4].value

    if str(data) > '2022-10-01 00:00:00' and baza == 'Сбербанк':

        print(a, b, c, data, baza)


        ws[r][0].value = sheet[row][0].value
        ws[r][1].value = sheet[row][1].value
        ws[r][2].value = sheet[row][2].value
        ws[r][3].value = sheet[row][3].value
        ws[r][4].value = sheet[row][4].value
        r += 1


wb.save('new file.xlsx')
wb.close()
print('ok')



